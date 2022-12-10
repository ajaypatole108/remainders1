from __future__ import unicode_literals
import frappe
import os
import frappe.utils
from frappe.utils import cint, cstr, flt, getdate, nowdate
from datetime import datetime
from frappe.desk import query_report
from six import string_types
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font,Alignment
from openpyxl.styles.borders import Border,Side
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from jinja2 import Environment, BaseLoader, PackageLoader,FileSystemLoader
# from email.MIMEBase import MIMEBase
import base64

__location__ = os.path.realpath(os.path.join(os.getcwd(), os.path.dirname(__file__)))
print('location: ',__location__)

@frappe.whitelist()
def generate_customer_outstanding_data(customer_name):
	print('\n\----------> Customer Name: ', customer_name)
	customer_filter = get_data(customer_name)
	print('\ncustomer_filter: ',customer_filter)

	thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

	if len(customer_filter) != 0 :
		wb = load_workbook(os.path.join(__location__, 'Outstanding_remainder_template.xlsx'))
		template_sheet = wb['Remainder']
		template_sheet['A14'] = customer_name

		for row in range(18,len(customer_filter)+18):
			i = 1
			for key in customer_filter[row-18].keys():

				# This code for applying border to cell
				for i1 in range(row,row+1):
					for j1 in range(1,11):
						ch = chr(64+j1)
						template_sheet[ch+(str(i1))].border = thin_border

				if key == 'name':
					_ = template_sheet.cell(column=1, row=row, value= customer_filter[row-18][key])
				elif key == 'posting_date':
					_ = template_sheet.cell(column=2, row=row, value= customer_filter[row-18][key])
				elif key == 'po_no':
					_ = template_sheet.cell(column=3, row=row, value= customer_filter[row-18][key])
				elif key == 'po_date':
					_ = template_sheet.cell(column=4, row=row, value= customer_filter[row-18][key])
				elif key == 'due_date':
					_ = template_sheet.cell(column=5, row=row, value= customer_filter[row-18][key])
				elif key == 'age':
					_ = template_sheet.cell(column=6, row=row, value= customer_filter[row-18][key])
				elif key == 'base_rounded_total':
					_ = template_sheet.cell(column=7, row=row, value= customer_filter[row-18][key])
				elif key == 'paid_amount':
					_ = template_sheet.cell(column=8, row=row, value= customer_filter[row-18][key])
				elif key == 'cn_amount':
					_ = template_sheet.cell(column=9, row=row, value= customer_filter[row-18][key])
				elif key == 'outstanding_amount':
					_ = template_sheet.cell(column=10, row=row, value= customer_filter[row-18][key])
					# template_sheet.column_dimensions['J'].width = 15
					_.font = Font(bold=True,size=10)

		#Blank Line
		cell1 = f"A{str(row+1)}:F{str(row+1)}"
		template_sheet.merge_cells(f'{cell1}')
		template_sheet.row_dimensions[row+1].height = 5

		template_sheet['A'+(str(row+2))] = "Grand Total"
		template_sheet['A'+(str(row+2))].font = Font(bold=True,size=10)
		template_sheet['A'+(str(row+2))].alignment = Alignment(horizontal='center')
		template_sheet['A'+(str(row+2))].border = thin_border
		cell1 = f"B{str(row+2)}:F{str(row+2)}"
		template_sheet.merge_cells(f'{cell1}')

		# Merge Cell Border
		template_sheet['B'+(str(row+2))].border = thin_border
		template_sheet['C'+(str(row+2))].border = thin_border
		template_sheet['D'+(str(row+2))].border = thin_border
		template_sheet['E'+(str(row+2))].border = thin_border
		template_sheet['F'+(str(row+2))].border = thin_border
		
		template_sheet['G'+(str(row+2))] = f"=SUM({get_column_letter(7)+'18'}:{get_column_letter(7)+str(row)})"
		template_sheet['G'+(str(row+2))].font = Font(bold=True,size=10)
		template_sheet['G'+(str(row+2))].border = thin_border
		sum_inv_amt = 0
		# Calculating Grand Total Of Column G
		for i in range(18,len(customer_filter)+18):
			a1 = template_sheet.cell(column=7,row=i).value
			if a1 != None:
				sum_inv_amt = sum_inv_amt + a1

		template_sheet['H'+(str(row+2))] = f"=SUM({get_column_letter(8)+'18'}:{get_column_letter(8)+str(row)})"
		template_sheet['H'+(str(row+2))].font = Font(bold=True,size=10)
		template_sheet['H'+(str(row+2))].border = thin_border
		sum_paid_amt = 0
		# Calculating Grand Total Of Column H
		for i in range(18,len(customer_filter)+18):
			a1 = template_sheet.cell(column=8,row=i).value
			if a1 != None:
				sum_paid_amt = sum_paid_amt + a1

		template_sheet['I'+(str(row+2))] = f"=SUM({get_column_letter(9)+'18'}:{get_column_letter(9)+str(row)})"
		template_sheet['I'+(str(row+2))].font = Font(bold=True,size=10)
		template_sheet['I'+(str(row+2))].border = thin_border
		sum_cn_amt = 0
		# Calculating Grand Total Of Column I
		for i in range(18,len(customer_filter)+18):
			a1 = template_sheet.cell(column=9,row=i).value
			if a1 != None:
				sum_cn_amt = sum_cn_amt + a1

		template_sheet['J'+(str(row+2))] = f"=SUM({get_column_letter(10)+'18'}:{get_column_letter(10)+str(row)})"
		template_sheet['J'+(str(row+2))].font = Font(bold=True,size=10)
		template_sheet['J'+(str(row+2))].border = thin_border
		sum_outstanding_amt = 0
		# Calculating Grand Total Of Column J
		for i in range(18,len(customer_filter)+18):
			a1 = template_sheet.cell(column=10,row=i).value
			if a1 != None:
				sum_outstanding_amt = sum_outstanding_amt + a1

		template_sheet.row_dimensions[row+3].height = 5

		wb.save(os.path.join(__location__, 'Outstanding_remainder_template_1.xlsx'))
		wb.close()
		# print('\n\nreturn: ',customer_filter,sum_inv_amt,sum_paid_amt,sum_cn_amt,sum_outstanding_amt)

		return customer_filter,sum_inv_amt,sum_paid_amt,sum_cn_amt,sum_outstanding_amt

def get_customer():
    data = frappe.db.sql(f"""
                            SELECT customer_name,email_id1 as email_id
                            FROM `tabCustomer`
                            """,as_dict=1)
    return data

def get_data(cust):
	# print(f"\n\n\n{cust}\n\n\n")
	data = frappe.db.sql(f"""
							SELECT UNIQUE si.name,si.posting_date,si.customer,si.po_no,si.po_date,si.due_date,si.base_rounded_total,si.outstanding_amount,
							(SELECT -(rounded_total) FROM `tabSales Invoice` cnsi WHERE cnsi.return_against = si.name) AS cn_amount
							FROM
							`tabSales Invoice` si
							LEFT JOIN
							`tabPayment Entry Reference` per
							ON
							si.name = per.reference_name
							WHERE
							si.customer = '{cust}'
							AND
							si.outstanding_amount > 0
							AND
							si.status != 'Draft' AND si.status != 'Cancelled'
							AND
							si.status != 'Paid'
							ORDER BY si.posting_date
							""",as_dict=1)
	# print('data: ',data)

	data1 = frappe.db.sql(f"""
							SELECT pe.name, pe.party_name as customer,pe.posting_date,pe.unallocated_amount as paid_amount
							FROM
							`tabPayment Entry` pe
							WHERE
							pe.party_name = '{cust}'
							AND
							pe.unallocated_amount != 0
							AND
							pe.status != 'Cancelled'
							ORDER BY pe.posting_date
						""",as_dict=1)
	# print('data1: ',data1)
 
	d = []
	if len(data) != 0:
		for i in data:
			doc1 = frappe.get_doc('Sales Invoice',i.name)
			age1 = (getdate(nowdate()) - getdate(doc1.posting_date)).days
			i['age'] = age1
			d.append(i)  # return dictionary {} but it want list so added into list d [{},{}]

	if len(data1) != 0:
		for i in data1:
			doc1 = frappe.get_doc('Payment Entry',i.name)
			i['outstanding_amount'] = -(doc1.unallocated_amount)
			d.append(i)
	# print('d: ',d)
	return d

@frappe.whitelist()
def send_outstanding_mail():
	customer = get_customer()
	print('customer: ',customer,'\n\n')
	if len(customer) != 0:
		for i in customer:
			ret_data = generate_customer_outstanding_data(i.customer_name)
			if ret_data != None :
				customer_outstanding_data,sum_inv_amt,sum_paid_amt,sum_cn_amt,sum_outstanding_amt = ret_data
				print('Customer return data: ',customer_outstanding_data,sum_inv_amt,sum_paid_amt,sum_cn_amt,sum_outstanding_amt)

				attachments = [{
						'fname': "Outstanding_remainder_template.xlsx",
						'fcontent': open(os.path.join(__location__, 'Outstanding_remainder_template_1.xlsx'),'rb').read()
					}]

				message = """
						Dear Sir/Madam,<br>
						<br>
						We would like to draw your kind attention at our following bill for payment, details as under:-
						<br>
						<br>
						"""

				cust = {'customer_outstanding':customer_outstanding_data,
						'customer_name': customer_outstanding_data[0].customer,
						'sum_inv_amt': sum_inv_amt,
						'sum_paid_amt': sum_paid_amt,
						'sum_cn_amt': sum_cn_amt,
						'sum_outstanding_amt': sum_outstanding_amt,
						}

				rtemplate = Environment(loader=FileSystemLoader(__location__))
				html_file = rtemplate.get_template('email_template.html')
				email_template = html_file.render(**cust)
				# print('email_template: ',email_template) # return html code 

				frappe.sendmail(
								# recipients = i.email_id,
								recipients = 'ajaypato.com@gmail.com',
								subject = "Outstanding Statement",
								message = message + email_template,
								attachments = attachments
							)
				frappe.db.commit()
				break

@frappe.whitelist()
def update_email_id(customer_name1,email_id):
	print('customer_name: ',customer_name1,email_id)

	frappe.db.set_value('Customer',customer_name1,'email_id1',email_id)